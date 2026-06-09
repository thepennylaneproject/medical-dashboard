#!/usr/bin/env python3
"""Export sarah_variants_formatted.xlsx → variants_clinical_formatted.json for the portal."""

import json
import sys
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
OUT = ROOT / "public/documents/variants_clinical_formatted.json"

try:
    import openpyxl
except ImportError:
    if OUT.exists():
        print(f"openpyxl not installed — using existing {OUT.name}")
        sys.exit(0)
    sys.exit("Install openpyxl: pip3 install openpyxl")

DEFAULT_XLSX = Path.home() / "Downloads" / "sarah_variants_formatted.xlsx"
XLSX = Path(sys.argv[1]) if len(sys.argv) > 1 else ROOT / "public/documents/sarah_variants_formatted.xlsx"
if not XLSX.exists():
    XLSX = DEFAULT_XLSX

CATEGORY_COLORS = {
    "AUTONOMIC & CARDIAC": "#831843",
    "BONE & MUSCULOSKELETAL": "#374151",
    "CONNECTIVE TISSUE": "#1F4E79",
    "HOMOCYSTEINE & METHYLATION": "#1E3A5F",
    "INFLAMMATORY CYTOKINES": "#7C2D12",
    "NEURODIVERGENCE & PSYCHIATRIC": "#4C1D95",
    "PHARMACOGENOMICS": "#1F4E79",
    "PLATELET & COAGULATION": "#78350F",
    "PULMONARY & ASTHMA": "#065F46",
}


def parse_clinical_sheet(ws):
    rows = []
    for r in range(5, ws.max_row + 1):
        label = ws.cell(r, 1).value
        if label is None:
            continue
        if isinstance(label, str) and label.startswith("  ") and ws.cell(r, 2).value is None:
            key = label.strip().upper()
            rows.append(
                {
                    "type": "category",
                    "label": label.strip(),
                    "color": CATEGORY_COLORS.get(key, "#1F4E79"),
                }
            )
            continue
        if label == "rsID":
            continue
        rows.append(
            {
                "type": "variant",
                "rsid": label,
                "gene": ws.cell(r, 2).value,
                "annotation": ws.cell(r, 3).value,
                "category": ws.cell(r, 4).value,
                "chrom": ws.cell(r, 5).value,
                "pos": ws.cell(r, 6).value,
                "genotype": ws.cell(r, 7).value,
                "zygosity": ws.cell(r, 8).value,
                "note": ws.cell(r, 9).value,
            }
        )
    return rows


def parse_table(ws, start_row):
    headers = [ws.cell(start_row, c).value for c in range(1, ws.max_column + 1)]
    rows = []
    for r in range(start_row + 1, ws.max_row + 1):
        vals = [ws.cell(r, c).value for c in range(1, ws.max_column + 1)]
        if all(v is None for v in vals):
            continue
        rows.append(dict(zip(headers, vals)))
    return {"headers": headers, "rows": rows}


def main():
    if not XLSX.exists():
        sys.exit(f"Missing workbook: {XLSX}")

    wb = openpyxl.load_workbook(XLSX, data_only=True)
    out = {
        "title": wb["Clinical Variants"]["A1"].value,
        "subtitle": wb["Clinical Variants"]["A2"].value,
        "clinicalVariants": parse_clinical_sheet(wb["Clinical Variants"]),
        "heterozygous": parse_table(wb["Heterozygous Variants"], 3),
        "pharmacogenomics": {
            "title": wb["Pharmacogenomics"]["A1"].value,
            "subtitle": wb["Pharmacogenomics"]["A2"].value,
            **parse_table(wb["Pharmacogenomics"], 3),
        },
    }

    OUT.parent.mkdir(parents=True, exist_ok=True)
    OUT.write_text(json.dumps(out, indent=2))
    variants = len([x for x in out["clinicalVariants"] if x["type"] == "variant"])
    print(f"Wrote {OUT} ({variants} clinical variants)")


if __name__ == "__main__":
    main()
